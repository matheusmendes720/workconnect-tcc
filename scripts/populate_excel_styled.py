import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

file_path = r'c:\Users\mathe\code_space\senai\workconnect-tcc\Matheus Mendes Conceição Santana Santos - PoC-Prova de conceito.xlsx'

demanda_text = "DEMANDA: (WorkConnect) Sistema de gestão inteligente de estoque para PMEs, com dashboards analíticos, conformidade LGPD e controles rigorosos de acesso."

# 7 requirements to fit the 2 + 3 + 2 layout perfectly
requisitos = [
    # Group 1: 2 items (A4:A5)
    ["Funcional", "Gestão de Produtos", "Permitir o cadastro, edição e exclusão de produtos com controle de quantidade em tempo real.", "Usuário consegue adicionar novo item e visualizar no dashboard.", "Alta", "Core do sistema."],
    ["Funcional", "Dashboard Analítico", "Exibir gráficos de tendências sazonais e saúde do estoque usando bibliotecas visuais.", "Gráficos renderizam corretamente com dados do banco ao acessar rota principal.", "Alta", "Tomada de decisão em PMEs."],
    
    # Group 2: 3 items (A7:A9)
    ["Segurança e Conformidade", "Autenticação Hashing", "Sistema de login com bcrypt para hashing longo de senhas.", "Usuário não autenticado é redirecionado; senhas protegidas.", "Alta", "Evita vazamento de dados."],
    ["Segurança e Conformidade", "LGPD - Exportação", "Módulo do usuário com aceite de termos de privacidade e exportação de dados JSON.", "Usuário consegue visualizar políticas e baixar dados estruturados.", "Alta", "Diferencial de conformidade."],
    ["Segurança e Conformidade", "Controle de Acessos", "Interface exibe funções limitadas com base no nível de privilégio do usuário autenticado.", "Painel adaptado de acordo com a regra de negócio do usuário.", "Média", "Segurança corporativa."],

    # Group 3: 2 items (A11:A12)
    ["Desempenho e UI", "Design Responsivo", "A interface web adapta-se a telas móveis (ex: 375px) usando classes Tailwind.", "Tabelas de estoque legíveis no mobile sem quebra horizontal extrema.", "Média", "Uso por gerentes em movimento."],
    ["Desempenho e UI", "Alertas de Estoque", "O sistema sinaliza visualmente na interface quando um item ultrapassa a quantidade mínima.", "Linhas apresentam ícones de emergência se a quantidade estiver crítica.", "Alta", "Evita desabastecimento."]
]

try:
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # Def styles
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=24, name="Calibri")
    bold_font = Font(bold=True, size=11, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Row 1
    sheet.merge_cells('A1:F1')
    c = sheet['A1']
    c.value = "PROVA DE CONCEITO"
    c.fill = blue_fill
    c.font = white_font
    c.alignment = center_align
    sheet.row_dimensions[1].height = 40
    
    # Row 2
    sheet.merge_cells('A2:F2')
    c2 = sheet['A2']
    c2.value = demanda_text
    c2.fill = yellow_fill
    c2.font = bold_font
    c2.alignment = left_align
    sheet.row_dimensions[2].height = 25
    
    # Row 3 (Headers)
    headers = ["Categoria", "Requisito", "Descrição Detalhada", "Critério de Aceite", "Prioridade", "Observações"]
    for col, h in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col)
        cell.value = h
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Define widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 25
    
    def apply_style(row_idx, req, cat=None):
        if cat:
            c = sheet.cell(row=row_idx, column=1)
            c.value = cat
            c.font = bold_font
            c.alignment = center_align
            c.border = thin_border
        
        for col in range(2, 7):
            c = sheet.cell(row=row_idx, column=col)
            c.value = req[col-1]
            c.font = normal_font
            c.alignment = center_align if col in [5] else left_align
            c.border = thin_border
            
    # Group 1 (Rows 4, 5)
    sheet.merge_cells('A4:A5')
    apply_style(4, requisitos[0], cat=requisitos[0][0])
    apply_style(5, requisitos[1])
    
    # Row 6 divider
    sheet.row_dimensions[6].height = 8
    
    # Group 2 (Rows 7, 8, 9)
    sheet.merge_cells('A7:A9')
    apply_style(7, requisitos[2], cat=requisitos[2][0])
    apply_style(8, requisitos[3])
    apply_style(9, requisitos[4])
    
    # Row 10 divider
    sheet.row_dimensions[10].height = 8

    # Group 3 (Rows 11, 12)
    sheet.merge_cells('A11:A12')
    apply_style(11, requisitos[5], cat=requisitos[5][0])
    apply_style(12, requisitos[6])
    
    # Apply borders to the merged category cells to make sure they look complete
    for r in [5, 8, 9, 12]:
        sheet.cell(row=r, column=1).border = thin_border

    wb.save(file_path)
    print("Beautifully styled Excel file successfully generated.")

except Exception as e:
    print(f"Failed to write to excel: {e}")
